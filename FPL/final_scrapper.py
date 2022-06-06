from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import pandas as pd
import numpy as np
from time import sleep, strftime
from random import randint
from openpyxl import load_workbook
import os


def main():
    # Page is preopened with OPTA Page loaded
    chrome_options = ChromeOptions()
    chrome_options.add_experimental_option("debuggerAddress", "localhost:1991")
    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(driver, 10)

    url = "https://www.fantasyfootballhub.co.uk/opta"
    slider_1 = driver.find_element(By.XPATH, "(//span)[18]")
    slider_2 = driver.find_element(By.XPATH, "(//span)[22]")

    click_per_start(driver)
    
    move_sliders_and_scrape(driver, slider_1, slider_2)


def move_sliders_and_scrape(driver, slider_1, slider_2):
    SLIDER_WIDTH = 565
    TOTAL_GWS = 38
    PREDICTION_RANGE = 3
    MAX = 8 # Which GW should the program stop scraping

    pixels_per_gw = SLIDER_WIDTH / (TOTAL_GWS - 1)
    gws_to_consider = 4

   # Reset the sliders. Lower slider and Upper slider to GW1
    ActionChains(driver).drag_and_drop_by_offset(slider_1, -SLIDER_WIDTH, 0).perform()
    short_sleep()
    ActionChains(driver).drag_and_drop_by_offset(slider_2, -SLIDER_WIDTH, 0).perform() 
    short_sleep()

    # Move upper slider to prediction GW
    ActionChains(driver).drag_and_drop_by_offset(slider_2, ((gws_to_consider - 1) * pixels_per_gw), 0).perform() # Then move the upper limit to GW3
    random_sleeps()

    which_gw_are_we_on = 1 # Initialise to 1

    while which_gw_are_we_on < MAX:
        data = scrape_players(driver)
        print("Getting the main player data...")
        
        short_sleep()

        # Move the sliders forward for the predict scrapping
        # Upper slider by PREDICTION RANGE and lower slider by GWs being considered
        ActionChains(driver).drag_and_drop_by_offset(slider_2, (pixels_per_gw * PREDICTION_RANGE), 0).perform()
        short_sleep()
        ActionChains(driver).drag_and_drop_by_offset(slider_1, (pixels_per_gw * gws_to_consider), 0).perform()
        random_sleeps()

        which_gw_are_we_on = int(slider_2.text)
        print(f"We are on GW {which_gw_are_we_on}")

        prediction = scrape_players_predict(driver)
        print("Getting prediction data...")

        # Merge all the dataFrames using the Name column to match the data across dataFrames 
        data_all = data.merge(prediction, how='left', on='Names')

        # Fill empty cells with NaN
        data_all.replace("", np.nan, inplace=True)

        # Remove all rows that contain NaN
        data_all.dropna(subset=["Points_y"], inplace=True)

        print("Merged all the data...")

        data_all.to_excel(f"Players with Predicted Points Scrape - {strftime('%a %d %b %Y %H %M %S %p')}.xlsx", index=False)

        short_sleep()

        # append_df_to_excel('202122 FWD Per App 2 GW.xlsx', data_all, index=False, header = None)
        data_all.to_csv("hello.csv", mode="a", index=False, header=False)

        short_sleep()

        # Return sliders to position before prediction moved the slider
        ActionChains(driver).drag_and_drop_by_offset(slider_1, -(pixels_per_gw * gws_to_consider), 0).perform()
        short_sleep()
        ActionChains(driver).drag_and_drop_by_offset(slider_2, -(pixels_per_gw * PREDICTION_RANGE), 0).perform()
        short_sleep()

        # Move the sliders by 1 GW each so the scraping process can be rinsed and repeated
        ActionChains(driver).drag_and_drop_by_offset(slider_2, (pixels_per_gw), 0).perform()
        short_sleep()
        ActionChains(driver).drag_and_drop_by_offset(slider_1, (pixels_per_gw), 0).perform()
        short_sleep()

def scrape_players_and_teams(driver, wait):

    # Read the page source
    html = driver.page_source

    # Create panda DataFrames for all possible tables on the HTML page
    df = pd.read_html(html)

    # For easy selection of the 1st table (In my case the only table)
    data = df[0]

    # Split the Name column into Name, Position and Team.
    # Has to be done in 2 different batches
    data[["Names", "Position"]] = data["Name"].str.split("(", expand=True)
    data[["Position", "Team"]] = data["Position"].str.split(")", expand=True)

    # Remove all instances of % and £ in the DataFrame
    data = data.replace(regex=["%"], value=[""])
    data = data.replace(regex=["£"], value=[""])

    random_sleeps()

    team_data_type(driver, wait)

    """Converts a team's long name to their short name."""
    team_name_conversion = {
        "Arsenal": "ARS",
        "Aston Villa": "AVL",
        "Brentford": "BRE",
        "Brighton": "BHA",
        "Burnley": "BUR",
        "Bournemouth": "BOU",
        "Cardiff": "CAR",
        "Chelsea": "CHE",
        "Crystal Palace": "CRY",
        "Everton": "EVE",
        "Fulham": "FUL",
        "Huddersfield": "HUD",
        "Leicester": "LEI",
        "Leeds": "LEE",
        "Liverpool": "LIV",
        "Man City": "MCI",
        "Man Utd": "MUN",
        "Newcastle": "NEW",
        "Norwich": "NOR",
        "Sheffield Utd": "SHU",
        "Southampton": "SOU",
        "Spurs": "TOT",
        "Watford": "WAT",
        "West Brom": "WBA",
        "West Ham": "WHU",
        "Wolves": "WOL",
        None: None,
    }

    team_html = driver.page_source

    dp = pd.read_html(team_html)

    teams = dp[0]

    teams.rename(columns={"Name": "Team"}, inplace=True)

    # Replace all the team names so they match up
    for word, replacement in team_name_conversion.items():
        teams = teams.replace(regex=[word], value=[replacement])

    combined = data.merge(teams, how="left", on="Team")

    # combined.to_excel(f"Players & Teams Combined Scrape - {strftime('%a %d %b %Y %H %M %S %p')}.xlsx", index=False)

    return combined


def scrape_players(driver):
    # Read the page source 
    html = driver.page_source

    # Create panda DataFrames for all possible tables on the HTML page
    df = pd.read_html(html)

    # For easy selection of the 1st table (In my case the only table)
    data = df[0]

    # Split the Name column into Name, Position and Team.
    # Has to be done in 2 different batches
    data[["Names", "Position"]] = data["Name"].str.split("(", expand=True)
    data[["Position", "Team"]] = data["Position"].str.split(")", expand=True)

    # Remove all instances of % and £ in the DataFrame
    data = data.replace(regex=["%"], value=[""])
    data = data.replace(regex=["£"], value=[""])

    # data.to_excel(f"Players Only Scrape - {strftime('%a %d %b %Y %H %M %S %p')}.xlsx", index=False)

    return data

def scrape_players_predict(driver):
    # Read the page source 
    html = driver.page_source

    # Create panda DataFrames for all possible tables on the HTML page
    df = pd.read_html(html)

    # For easy selection of the 1st table (In my case the only table)
    data_predict = df[0]

    # Split the Name column into Name, Position and Team.
    # Has to be done in 2 different batches
    data_predict[["Names", "Position"]] = data_predict["Name"].str.split("(", expand=True)

    # Drop all columns except Names and Points
    data_predict.drop(data_predict.columns.difference(["Names", "Points"]), 1, inplace=True)

    # data_predict.to_excel(f"Prediction Only Scrape - {strftime('%a %d %b %Y %H %M %S %p')}.xlsx", index=False)

    return data_predict


def maximize_window(driver):
    driver.maximize_window()

    random_sleeps()


def load_all_players(driver, wait):
    driver.find_element(By.CSS_SELECTOR, "div[class='my-4'] div[class=' css-r71ql9-singleValue']").click()

    # Load all possible players
    load_all = wait.until(
        EC.visibility_of_element_located(
            (By.XPATH, "//div[contains(text(),'All (slow)')]")
        )
    )

    load_all.click()

    sleep(20)

def click_per90(driver):
    driver.find_element(By.CSS_SELECTOR, "#per90").click()

    random_sleeps()

def click_per_start(driver):
    driver.find_element(By.CSS_SELECTOR, "#perstart").click()

    random_sleeps()

def click_perapp(driver):
    driver.find_element(By.CSS_SELECTOR, "#perapp").click()
    
    random_sleeps()

def team_data_type(driver, wait):
    # this will click dropdown for data type
    try: 
        driver.find_element(
            By.XPATH, value="//div[contains(text(),'Players')]"
        ).click()

    except:
        driver.find_element(By.XPATH, value="//div[contains(text(),'Teams')]").click()  

    # Click on Player option
    teams = wait.until(
        EC.visibility_of_element_located((By.XPATH, "//div[contains(text(),'Teams')]"))
    )

    teams.click()

    random_sleeps()

def player_data_type(driver, wait):
    # this will click dropdown for data type
    driver.find_element(
        By.XPATH, value="(//div)[132]"
    ).click()

    # Click on Player option
    player = wait.until(
        EC.visibility_of_element_located(
            (By.XPATH, "//div[contains(text(),'Players')]")
        )
    )

    player.click()

    random_sleeps()


def stat_type_custom(driver, wait):
    """
    This function goes into the Stat Type option and selects Custom
    """
    # this will click dropdown for stat types
    driver.find_element(
        By.XPATH, value="(//div[contains(@class,'css-seq4h5-control')])[3]"
    ).click()

    random_sleeps()

    # Click on Custom option
    custom = wait.until(
        EC.visibility_of_element_located((By.XPATH, "//div[contains(text(),'Custom')]"))
    )

    custom.click()

    random_sleeps()

    # Click Save button for the custom selections
    driver.find_element(By.XPATH, value="//button[normalize-space()='Save']").click()

    random_sleeps()

def select_fwd_position(driver):
    # Select FWD - Requires unclicking all the other positions (GK, DEF, MID)
    for i in {0, 1, 2}:
        unclick_positions(driver, i)
    
    random_sleeps()

def unclick_positions(driver, i):
    driver.find_element(By.XPATH, "//input[@id=" + str(i) + "]").click()

def select_mid_position(driver):
    # Select MID - Requires unclicking all the other positions (GK, DEF, FWD)
    for i in {0, 1, 3}:
        unclick_positions(driver, i)

    random_sleeps()

def select_def_position(driver):
    # Select DEF - Requires unclicking all the other positions (GK, MID, FWD)
    for i in {0, 2, 3}:
        unclick_positions(driver, i)
    
    random_sleeps()

def select_gk_position(driver):
    # Select GK - Requires unclicking all the other positions (DEF, MID, FWD)
    for i in range(1, 4):
        unclick_positions(driver, i)
    
    random_sleeps()

def random_sleeps():
    sleep(randint(15, 20))

def short_sleep():
    sleep(randint(3, 6))

def season_202122(driver, wait):
    click_seasons_box(driver)

    year_2021 = wait.until(
        EC.visibility_of_element_located((By.XPATH, "//div[contains(text(),'2021/22')]"))
    )

    year_2021.click()

    random_sleeps()

def click_seasons_box(driver):
    driver.find_element(By.XPATH, "(//div[contains(@class,'css-seq4h5-control')])[2]").click()

def season_202021(driver, wait):
    click_seasons_box(driver)

    year_2020 = wait.until(
        EC.visibility_of_element_located((By.XPATH, "//div[contains(text(),'2020/21')]"))
    )

    year_2020.click()
    
    random_sleeps()

def season_201920(driver, wait):
    click_seasons_box(driver)

    year_2019 = wait.until(
        EC.visibility_of_element_located((By.XPATH, "//div[contains(text(),'2019/20')]"))
    )

    year_2019.click()
    
    random_sleeps()

def season_201819(driver, wait):
    click_seasons_box(driver)

    year_2018 = wait.until(
        EC.visibility_of_element_located((By.XPATH, "//div[contains(text(),'2018/19')]"))
    )

    year_2018.click()
    
    random_sleeps()

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists = "overlay")

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

if __name__ == "__main__":
    main()
