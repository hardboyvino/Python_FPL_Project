from matplotlib.pyplot import draw_if_interactive
import numpy as np
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import pandas as pd
from time import sleep, strftime
from random import randint
from openpyxl import load_workbook


def main():
    # Page is preopened with OPTA Page loaded
    chrome_options = ChromeOptions()
    chrome_options.add_experimental_option("debuggerAddress", "localhost:1991")
    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(driver, 10)

    SLIDER_WIDTH = 570
    TOTAL_GWS = 38
    PREDICTION_RANGE = 3
    MAX = 7

    pixels_per_gw = SLIDER_WIDTH / (TOTAL_GWS - 1)
    gws_to_consider = 2
    slider_1 = driver.find_element(By.XPATH, "(//span)[18]")
    slider_2 = driver.find_element(By.XPATH, "(//span)[22]")
    
    move_sliders_and_scrape(driver, SLIDER_WIDTH, PREDICTION_RANGE, MAX, pixels_per_gw, gws_to_consider, slider_1, slider_2)

def move_sliders_and_scrape(driver, SLIDER_WIDTH, PREDICTION_RANGE, MAX, pixels_per_gw, gws_to_consider, slider_1, slider_2):
    # Reset the sliders. Lower slider and Upper slider to GW1
    ActionChains(driver).drag_and_drop_by_offset(slider_1, -SLIDER_WIDTH, 0).perform()
    random_sleeps()
    ActionChains(driver).drag_and_drop_by_offset(slider_2, -SLIDER_WIDTH, 0).perform() 
    random_sleeps()

    # Move upper slider to prediction GW
    ActionChains(driver).drag_and_drop_by_offset(slider_2, ((gws_to_consider - 1) * pixels_per_gw), 0).perform() # Then move the upper limit to GW3
    random_sleeps()

    which_gw_are_we_on = 1 # Initialise to 1

    while which_gw_are_we_on < MAX:
        data = scrape_players(driver)
        print(data)
        
        random_sleeps()

        # Move the sliders forward for the predict scrapping
        # Upper slider by PREDICTION RANGE and lower slider by GWs being considered
        ActionChains(driver).drag_and_drop_by_offset(slider_2, (pixels_per_gw * PREDICTION_RANGE), 0).perform()
        random_sleeps()
        ActionChains(driver).drag_and_drop_by_offset(slider_1, (pixels_per_gw * gws_to_consider), 0).perform()
        random_sleeps()

        which_gw_are_we_on = int(slider_2.text)
        print(which_gw_are_we_on)

        prediction = scrape_players_predict(driver)
        print(prediction)

        # Merge all the dataFrames using the Name column to match the data across dataFrames 
        data_all = data.merge(prediction, how='left', on='Names')

        # # Fill all blanks in the dataframe with zeros so the results don't return NaN
        # data_all.fillna(0, inplace=True)

        # Fill empty cells with NaN
        data_all.replace("", np.nan, inplace=True)

        # Remove all rows that contain NaN
        data_all.dropna(subset=["Points_y"], inplace=True)

        print(data_all)

        data_all.to_excel(f"Players with Predicted Points Scrape - {strftime('%a %d %b %Y %H %M %S %p')}.xlsx", index=False)

        random_sleeps()

        book = load_workbook("final_countdown.xlsx")
        writer = pd.ExcelWriter("final_countdown.xlsx", engine="openpyxl")
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}

        for sheetname in writer.sheets:
            data_all.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False, header=True)

        writer.save()

        random_sleeps()

        # Return sliders to position before prediction moved the slider
        ActionChains(driver).drag_and_drop_by_offset(slider_1, -(pixels_per_gw * gws_to_consider), 0).perform()
        random_sleeps()
        ActionChains(driver).drag_and_drop_by_offset(slider_2, -(pixels_per_gw * PREDICTION_RANGE), 0).perform()
        random_sleeps()

        # Move the sliders by 1 GW each so the scraping process can be rinsed and repeated
        ActionChains(driver).drag_and_drop_by_offset(slider_2, (pixels_per_gw), 0).perform()
        random_sleeps()
        ActionChains(driver).drag_and_drop_by_offset(slider_1, (pixels_per_gw), 0).perform()
        random_sleeps()

def random_sleeps():
    sleep(randint(10, 13))

def scrape_players(driver):
    # Read the page source 
    html = driver.page_source

    # Create panda DataFrames for all possible tables on the HTML page
    df = pd.read_html(html)

    # For easy selection of the 1st table (In my case the only table)
    data = df[0]

    # Split the Name column into Name, Position and Team.
    # Has to be done in 2 different batches
    data[["Names", "Position"]] = data["Name"].str.split("(", expand=True)
    data[["Position", "Team"]] = data["Position"].str.split(")", expand=True)

    # Do not show the first column when print the DataFrame on VSCode, although the column still exists when exported to Excel or CSV
    # data = data.iloc[:, 1:]

    # Instead of just hiding the column, totally drop it so it is not even in the CSV or Excel
    # data = data.drop(columns=data.columns[0], axis=1, inplace=True)

    # Remove all instances of % and £ in the DataFrame
    data = data.replace(regex=["%"], value=[""])
    data = data.replace(regex=["£"], value=[""])

    # data.to_excel(f"Players Only Scrape - {strftime('%a %d %b %Y %H %M %S %p')}.xlsx", index=False)

    return data

def scrape_players_predict(driver):
    # Read the page source 
    html = driver.page_source

    # Create panda DataFrames for all possible tables on the HTML page
    df = pd.read_html(html)

    # For easy selection of the 1st table (In my case the only table)
    data_predict = df[0]

    # Split the Name column into Name, Position and Team.
    # Has to be done in 2 different batches
    data_predict[["Names", "Position"]] = data_predict["Name"].str.split("(", expand=True)

    # Drop all columns except Names and Points
    data_predict.drop(data_predict.columns.difference(["Names", "Points"]), 1, inplace=True)

    # data_predict.to_excel(f"Prediction Only Scrape - {strftime('%a %d %b %Y %H %M %S %p')}.xlsx", index=False)

    return data_predict

if __name__ == "__main__":
    main()